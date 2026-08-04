from __future__ import annotations

import pytest

from openpyxl import load_workbook

from src.application.audit_service import AuditRequest, AuditService
from src.application.progress import AuditCancelled
import main


class StateReporter:
    def __init__(self):
        self.state = {}

    def emit(self, _event, _message="", **data):
        self.state.update(data)


def test_audit_request_reuses_existing_date_and_region_validation():
    request = AuditRequest(
        start_date="2026-07-08",
        end_date="2026-07-09",
        regions=["충청북도", "충청남도"],
    )
    start, end = request.validated_dates()
    assert start.isoformat() == "2026-07-08"
    assert end.isoformat() == "2026-07-09"


@pytest.mark.parametrize(
    "audit_request",
    [
        AuditRequest("2026-07-09", "2026-07-08", ["충청북도"]),
        AuditRequest("2026-07-08", "2026-07-08", []),
        AuditRequest("2026-07-08", "2026-07-08", ["충청도"]),
        AuditRequest("2026-07-08", "2026-07-08", ["충청북도"], retries=3),
    ],
)
def test_audit_request_rejects_invalid_options(audit_request):
    with pytest.raises(ValueError):
        audit_request.validated_dates()


def test_cancelled_service_still_writes_existing_excel_structure(tmp_path, monkeypatch):
    async def cancelled_run(_collector):
        raise AuditCancelled("test cancellation")

    monkeypatch.setattr("src.application.audit_service.RegionalCollector.run", cancelled_run)
    result = AuditService(output_dir=tmp_path / "output", work_dir=tmp_path / "work").run(
        AuditRequest(
            start_date="2026-07-08",
            end_date="2026-07-08",
            regions=["충청북도"],
            checkpoint_path=tmp_path / "work" / "cancelled.jsonl",
        )
    )
    assert result.status == "cancelled"
    assert result.excel_path and result.excel_path.is_file()
    workbook = load_workbook(result.excel_path, read_only=True)
    assert workbook.sheetnames == ["점검결과", "오류목록", "점검요약", "디버그로그"]
    workbook.close()


def test_cancelled_service_preserves_last_operation_diagnostics(tmp_path, monkeypatch):
    async def cancelled_run(collector):
        collector.progress_reporter.emit(
            "link_started",
            current_operation="link_check:article-1",
            operation_started_at="2026-08-04T09:00:00+09:00",
            browser_state="running",
        )
        raise AuditCancelled("test cancellation")

    monkeypatch.setattr("src.application.audit_service.RegionalCollector.run", cancelled_run)
    reporter = StateReporter()
    result = AuditService(output_dir=tmp_path / "output", work_dir=tmp_path / "work").run(
        AuditRequest(
            start_date="2026-07-08",
            end_date="2026-07-08",
            regions=["충청북도"],
            checkpoint_path=tmp_path / "work" / "cancelled-diagnostics.jsonl",
        ),
        reporter=reporter,
    )

    assert result.status == "cancelled"
    assert reporter.state["current_operation"] == "link_check:article-1"
    assert reporter.state["operation_started_at"] == "2026-08-04T09:00:00+09:00"
    assert reporter.state["browser_state"] == "running"


def test_completed_service_clears_current_operation(tmp_path, monkeypatch):
    async def completed_run(collector):
        collector.progress_reporter.emit(
            "link_started",
            current_operation="link_check:article-1",
            operation_started_at="2026-08-04T09:00:00+09:00",
        )

    monkeypatch.setattr("src.application.audit_service.RegionalCollector.run", completed_run)
    reporter = StateReporter()
    result = AuditService(output_dir=tmp_path / "output", work_dir=tmp_path / "work").run(
        AuditRequest(
            start_date="2026-07-08",
            end_date="2026-07-08",
            regions=["충청북도"],
            checkpoint_path=tmp_path / "work" / "completed-diagnostics.jsonl",
        ),
        reporter=reporter,
    )

    assert result.status == "completed"
    assert reporter.state["current_operation"] == ""
    assert reporter.state["operation_started_at"] == ""


def test_cli_calls_common_audit_service(monkeypatch, tmp_path):
    captured = {}

    class FakeService:
        def run(self, request):
            captured["request"] = request
            return type("Result", (), {"status": "completed", "excel_path": tmp_path / "report.xlsx"})()

    monkeypatch.setattr(main, "AuditService", FakeService)
    exit_code = main.main([
        "--start-date", "2026-07-08",
        "--end-date", "2026-07-08",
        "--regions", "충청북도,충청남도",
        "--timeout", "20",
        "--retries", "1",
    ])
    assert exit_code == 0
    assert captured["request"].regions == ["충청북도", "충청남도"]
    assert captured["request"].timeout_seconds == 20
    assert captured["request"].retries == 1
