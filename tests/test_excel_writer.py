from datetime import datetime, timedelta

from openpyxl import load_workbook

from conftest import make_row
from src.config import DEBUG_COLUMNS, RESULT_COLUMNS
from src.excel_writer import write_excel
from src.logging_utils import debug_entry


def test_excel_has_four_sheets_columns_filters_and_links(tmp_path):
    rows = [
        make_row(),
        make_row(original_url="https://example.com/b", final_url="https://example.com/b", verdict="링크오류",
                 browser_result="표시 실패", link_working_yn="N", error_message="기사 페이지를 찾을 수 없음", source_order=2),
    ]
    started = datetime.now().astimezone()
    path = write_excel(tmp_path / "report.xlsx", rows, [], start_date="2026-07-01", end_date="2026-07-01",
                       started_at=started, ended_at=started + timedelta(seconds=3))
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["점검결과", "오류목록", "점검요약", "디버그로그"]
    result = workbook["점검결과"]
    errors = workbook["오류목록"]
    assert [cell.value for cell in result[1]] == RESULT_COLUMNS
    assert "링크작동여부_YN" in [cell.value for cell in errors[1]]
    assert result["M2"].hyperlink.target == "https://example.com/a"
    assert result["Q2"].value == "Y" and result["Q3"].value == "N"
    assert errors.max_row == 2 and errors["Q2"].value == "N"
    assert result.auto_filter.ref and errors.auto_filter.ref
    workbook.close()


def test_excel_summary_records_selected_region_information(tmp_path):
    rows = [make_row(region="서울특별시")]
    started = datetime.now().astimezone()
    path = write_excel(
        tmp_path / "selected.xlsx", rows, [], start_date="2026-07-01", end_date="2026-07-01",
        started_at=started, ended_at=started + timedelta(seconds=1),
        selected_regions=["서울특별시", "부산광역시", "경기도"],
    )
    workbook = load_workbook(path, read_only=True)
    summary = workbook["점검요약"]
    values = {row[0].value: row[1].value for row in summary.iter_rows(min_row=2, max_row=16)}
    assert values["지역 선택 방식"] == "선택"
    assert values["선택 지역 수"] == 3
    assert values["선택 지역 목록"] == "서울특별시, 부산광역시, 경기도"
    workbook.close()


def test_debug_sheet_separates_href_click_and_inferred_url_evidence(tmp_path):
    started = datetime.now().astimezone()
    entry = debug_entry(
        "링크URL",
        source_href_raw="www.example.com/a",
        source_href_property="https://www.bigkinds.or.kr/regional/www.example.com/a",
        click_target_raw="www.example.com/a",
        normalization_input="www.example.com/a",
        original_url="https://www.bigkinds.or.kr/regional/www.example.com/a",
        normalization_method="실제 클릭(새 탭)",
        click_before_url="https://www.bigkinds.or.kr/regional/curation.do",
        click_after_url="https://www.bigkinds.or.kr/regional/www.example.com/a",
        first_opened_url="https://www.bigkinds.or.kr/regional/www.example.com/a",
        new_tab_yn="Y", current_tab_moved_yn="N",
        inferred_url="https://www.example.com/a",
        url_structure_anomaly_yn="Y",
        url_structure_anomaly_details="BigKinds 경로 내부에 외부 도메인 문자열이 포함됨",
        final_url="https://www.bigkinds.or.kr/regional/www.example.com/a",
        http_status=404,
        access_reason_code="ACCESS_TEXT_MATCH",
        detected_phrase="접근 권한이 없습니다.",
        detected_locator="header .login-menu",
        detected_dom_area="header",
        detected_visible_yn="Y",
        document_title="접근 제한",
        visible_h1="접근 제한",
        article_exists_yn="N",
        primary_text_length=18,
        article_title_match_yn="N",
        article_rendered_yn="N",
    )
    path = write_excel(
        tmp_path / "debug.xlsx", [make_row()], [entry],
        start_date="2026-07-01", end_date="2026-07-01",
        started_at=started, ended_at=started + timedelta(seconds=1),
    )
    workbook = load_workbook(path, read_only=True)
    sheet = workbook["디버그로그"]
    assert [cell.value for cell in sheet[1]] == DEBUG_COLUMNS
    values = dict(zip(DEBUG_COLUMNS, [cell.value for cell in sheet[2]]))
    assert values["href속성원문"] == "www.example.com/a"
    assert values["href프로퍼티값"] == "https://www.bigkinds.or.kr/regional/www.example.com/a"
    assert values["클릭대상URL원문"] == "www.example.com/a"
    assert values["URL처리전"] == "www.example.com/a"
    assert values["원본URL"] == "https://www.bigkinds.or.kr/regional/www.example.com/a"
    assert values["URL처리방식"] == "실제 클릭(새 탭)"
    assert values["추정정상URL"] == "https://www.example.com/a"
    assert values["URL구조이상여부"] == "Y"
    assert values["HTTP상태"] == 404
    assert values["최종URL"] == "https://www.bigkinds.or.kr/regional/www.example.com/a"
    assert values["접근제한판정근거코드"] == "ACCESS_TEXT_MATCH"
    assert values["감지문구"] == "접근 권한이 없습니다."
    assert values["감지문구Locator"] == "header .login-menu"
    assert values["감지문구DOM영역"] == "header"
    assert values["감지문구Visible_YN"] == "Y"
    assert values["document.title"] == "접근 제한"
    assert values["visible h1"] == "접근 제한"
    assert values["article존재여부_YN"] == "N"
    assert values["주요콘텐츠텍스트길이"] == 18
    assert values["기사제목일치여부_YN"] == "N"
    workbook.close()
