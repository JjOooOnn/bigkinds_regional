from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import DEBUG_COLUMNS, RESULT_COLUMNS, TARGET_URL, VERDICTS
from .models import AuditRow, DebugEntry
from .summary import calculate_summary
from .regions import is_all_regions

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")
WARNING_FILL = PatternFill("solid", fgColor="FFEB9C")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")


def _result_values(index: int, row: AuditRow) -> list:
    return [
        index, row.requested_date, row.displayed_date, row.region, row.issue_order, row.issue_title,
        row.issue_categories, row.source_count, row.source_type, row.publisher, row.article_date,
        row.article_title, row.original_url, row.final_url, row.http_status, row.browser_result,
        row.link_working_yn, row.verdict, row.response_seconds, row.error_message, row.checked_at,
    ]


def _debug_values(entry: DebugEntry) -> list:
    return [
        entry.timestamp, entry.stage, entry.requested_date, entry.displayed_date, entry.region,
        entry.issue_order, entry.issue_title, entry.source_href_raw, entry.source_href_property,
        entry.click_target_raw, entry.normalization_input, entry.original_url,
        entry.normalization_method, entry.click_before_url, entry.click_after_url,
        entry.first_opened_url, entry.new_tab_yn, entry.current_tab_moved_yn,
        entry.inferred_url, entry.url_structure_anomaly_yn,
        entry.url_structure_anomaly_details, entry.locator, entry.event,
        entry.exception_type, entry.details, entry.http_status, entry.final_url, entry.screenshot_path,
        entry.retry_count,
        entry.access_reason_code, entry.detected_phrase, entry.detected_locator,
        entry.detected_dom_area, entry.detected_visible_yn, entry.document_title,
        entry.visible_h1, entry.article_exists_yn, entry.primary_text_length,
        entry.article_title_match_yn, entry.article_rendered_yn,
    ]


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _format_sheet(ws, header_row: int = 1) -> None:
    ws.freeze_panes = "A2"
    _style_header(ws, header_row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    preferred = {
        "이슈제목": 42, "기사제목": 52, "원본URL": 45, "최종URL": 45, "오류내용": 35,
        "상세내용": 55, "DOM또는Locator": 38, "스크린샷경로": 35, "지역명": 18,
        "점검일시": 25, "로그시각": 25, "href속성원문": 45, "href프로퍼티값": 45,
        "클릭대상URL원문": 45, "클릭직전URL": 45, "클릭직후URL": 45,
        "새페이지최초URL": 45, "추정정상URL": 45, "URL구조이상내용": 42,
        "접근제한판정근거코드": 38, "감지문구": 40, "감지문구Locator": 48,
        "감지문구DOM영역": 45, "document.title": 55, "visible h1": 50,
    }
    for index, column_cells in enumerate(ws.columns, 1):
        values = [str(cell.value or "") for cell in column_cells[:50]]
        header = str(ws.cell(header_row, index).value or "")
        width = preferred.get(header, min(max([len(value) for value in values] + [8]) + 2, 28))
        ws.column_dimensions[get_column_letter(index)].width = width


def _write_result_sheet(ws, rows: list[AuditRow]) -> None:
    ws.append(RESULT_COLUMNS)
    for index, row in enumerate(rows, 1):
        ws.append(_result_values(index, row))
    ws.auto_filter.ref = f"A1:U{max(ws.max_row, 1)}"
    _format_sheet(ws)
    for row_no in range(2, ws.max_row + 1):
        for col in (13, 14):
            cell = ws.cell(row_no, col)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
        ws.cell(row_no, 19).number_format = "0.000"
    if ws.max_row >= 2:
        ws.conditional_formatting.add(f"Q2:Q{ws.max_row}", CellIsRule(operator="equal", formula=['"N"'], fill=ERROR_FILL))
        ws.conditional_formatting.add(f"R2:R{ws.max_row}", CellIsRule(operator="equal", formula=['"정상"'], fill=GOOD_FILL))
        ws.conditional_formatting.add(f"R2:R{ws.max_row}", FormulaRule(formula=["R2<>\"정상\""], fill=ERROR_FILL))


def write_excel(
    path: Path, rows: list[AuditRow], debug_entries: list[DebugEntry],
    *, start_date: str, end_date: str, started_at: datetime, ended_at: datetime,
    region_count: int | None = None, issue_count: int | None = None,
    selected_regions: list[str] | None = None,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = sorted(rows, key=lambda r: (r.requested_date, r.region_order, r.issue_order, r.source_order))
    workbook = Workbook()
    result_ws = workbook.active
    result_ws.title = "점검결과"
    _write_result_sheet(result_ws, rows)

    error_ws = workbook.create_sheet("오류목록")
    _write_result_sheet(error_ws, [row for row in rows if row.verdict != "정상"])

    summary = calculate_summary(rows)
    summary_ws = workbook.create_sheet("점검요약")
    summary_ws.append(["항목", "값"])
    if selected_regions is None:
        selected_regions = list(dict.fromkeys(row.region for row in rows))
    else:
        # 호출자가 지정한 선택 순서를 Excel 실행정보에도 그대로 유지한다.
        selected_regions = list(selected_regions)
    all_regions = is_all_regions(selected_regions)
    info = [
        ("대상 URL", TARGET_URL), ("시작일", start_date), ("종료일", end_date),
        ("실행 시작시각", started_at.astimezone().isoformat(timespec="seconds")),
        ("실행 종료시각", ended_at.astimezone().isoformat(timespec="seconds")),
        ("소요시간", str(ended_at - started_at).split(".")[0]),
        ("지역 선택 방식", "전체" if all_regions else "선택"),
        ("선택 지역 수", len(selected_regions)),
        ("선택 지역 목록", "전체" if all_regions else ", ".join(selected_regions)),
        ("점검 지역 수", region_count if region_count is not None else len({(r.requested_date, r.region) for r in rows})),
        ("전체 이슈 수", issue_count if issue_count is not None else len({(r.requested_date, r.region, r.issue_order) for r in rows})),
        ("전체 링크 수", summary["total"]), ("정상 수", summary["normal"]),
        ("오류 수", summary["errors"]), ("전체 정상률", summary["rate"]),
    ]
    for item in info:
        summary_ws.append(item)
    summary_ws.cell(summary_ws.max_row, 2).number_format = "0.00%"
    table_row = summary_ws.max_row + 2
    summary_ws.append(["조회요청일", "지역명", "이슈수", "뉴스링크수", "정상수", "오류수", "정상률"])
    for detail in summary["details"]:
        summary_ws.append([detail["requested_date"], detail["region"], detail["issues"], detail["links"], detail["normal"], detail["errors"], detail["rate"]])
        summary_ws.cell(summary_ws.max_row, 7).number_format = "0.00%"
    summary_ws.auto_filter.ref = f"A{table_row}:G{max(summary_ws.max_row, table_row)}"
    verdict_row = summary_ws.max_row + 2
    summary_ws.append(["최종판정", "건수"])
    for verdict in VERDICTS:
        summary_ws.append([verdict, summary["verdict_counts"].get(verdict, 0)])
    _format_sheet(summary_ws)
    _style_header(summary_ws, table_row)
    _style_header(summary_ws, verdict_row)

    debug_ws = workbook.create_sheet("디버그로그")
    debug_ws.append(DEBUG_COLUMNS)
    for entry in debug_entries:
        debug_ws.append(_debug_values(entry))
    debug_ws.auto_filter.ref = f"A1:{get_column_letter(len(DEBUG_COLUMNS))}{max(debug_ws.max_row, 1)}"
    _format_sheet(debug_ws)

    workbook.save(path)
    # 저장 직후 다시 열어 손상 여부를 확인한다.
    check = load_workbook(path, read_only=True)
    check.close()
    return path.resolve()
