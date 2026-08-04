from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.checkpoint import CheckpointStore

from .job_repository import JobRepository


@dataclass(frozen=True)
class ArtifactCountComparison:
    job_id: str
    processed_links: int
    checkpoint_rows: int
    sqlite_rows: int
    excel_rows: int

    @property
    def matches(self) -> bool:
        return len({
            self.processed_links,
            self.checkpoint_rows,
            self.sqlite_rows,
            self.excel_rows,
        }) == 1

    def to_dict(self) -> dict[str, int | str | bool]:
        return {**asdict(self), "matches": self.matches}


def _excel_result_row_count(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "점검결과" not in workbook.sheetnames:
            raise ValueError("Excel 파일에 점검결과 시트가 없습니다.")
        worksheet = workbook["점검결과"]
        return sum(
            any(value is not None for value in row)
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        )
    finally:
        workbook.close()


def compare_job_artifact_counts(
    repository: JobRepository, job_id: str,
) -> ArtifactCountComparison:
    """한 작업의 DB·체크포인트·Excel 결과 행 수를 대조한다."""
    job = repository.get_job(job_id)
    if not job:
        raise KeyError(f"작업을 찾을 수 없습니다: {job_id}")

    checkpoint_path = Path(job["checkpoint_path"])
    if not checkpoint_path.is_file():
        raise FileNotFoundError(f"체크포인트 파일이 없습니다: {checkpoint_path}")
    excel_path = Path(job["excel_path"])
    if not job["excel_path"] or not excel_path.is_file():
        raise FileNotFoundError(f"Excel 파일이 없습니다: {excel_path}")

    checkpoint = CheckpointStore(checkpoint_path, resume=True)
    sqlite_rows = repository.get_result_summary(job_id)["total_links"]
    return ArtifactCountComparison(
        job_id=job_id,
        processed_links=int(job["processed_links"]),
        checkpoint_rows=len(checkpoint.rows),
        sqlite_rows=int(sqlite_rows),
        excel_rows=_excel_result_row_count(excel_path),
    )
